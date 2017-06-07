from TournamentClass import Tournament
from Scheduler import generateplayingfield, createpdfs, doscheduling, createscoresheets, createmasters
import time
import PlayerClass
import copy
from openpyxl import load_workbook, Workbook
from interval import interval




def main():

    # initialize a new tournament
    nhf = Tournament()
    originalfield = generateplayingfield("Final_1.xlsx", nhf)

    """ Fixes errors found in Final_1 """''
    wb = load_workbook("TOSENDTOERIC.XLSX")
    ws = wb.active
    names = []
    ids = []
    for row in ws:
        names.append(str(row[0].value) + " " + str(row[1].value))
        ids.append(str(row[4].value))
    ids = dict(zip(names, ids))
    for k in ids:
        for player in originalfield:
            if player.name == k:
                player.id = ids[k]
    wb.close()


    wb = load_workbook("Totals.xlsx", read_only=True,  data_only=True)
    ws = wb.active
    names = []
    ids = []
    sentplayernames = []
    needtosendplayers = []
    for row in ws:
        names.append(str(row[0].value) + " " + str(row[1].value))
        ids.append(str(row[4].value))
        sentplayernames.append(str(row[0].value) + " " + str(row[1].value))
    ids = dict(zip(names, ids))
    for k in ids:
        for player in originalfield:
            if player.name == k:
                player.id = ids[k]
    for player in originalfield:
        if player.name not in sentplayernames:
            needtosendplayers.append(player)
    wb.close()

    wb = Workbook()
    ws = wb.active
    ws.append(["First", "Last", "Division", "Seed", "ID", "Hometown"])
    count = 0
    for player in needtosendplayers:
        if "Extra" in player.name:
            continue
        name = player.name.split(" ")
        first = name[0]
        last = " ".join(name[1:])
        ws.append([first, last, player.division, str(player.seed).upper(), player.id, player.hometown])
        count += 1
    wb.save("TOSENDTOERIC.XLSX")
    wb.close()

    ids = []
    for player in originalfield:
        ids.append(player.id)
    if len(ids) != len(set(ids)):
        raise RuntimeError
    """ End fixing errors to Final_1 """


    count = 0
    for player in originalfield:
        for ev in player.schedule:
            if "History Bee Exam" == ev[0]:
                count += 1




    for player in originalfield:
        if player.name == "Andres Soto":
            print "Found 1"
            player.restriction.append(interval([100, 111.5]))
        if player.name == "Timothy Kashani":
            print "Found 2"
            player.restriction.append(interval([217.5, 220]))
        if player.name == "William Humble":
            print "Found 3"
            player.restriction.append(interval([100, 111.5]))
        if player.name == "Ashna Karia":
            print "Found 4"
            player.restriction.append(interval([100, 111.5]))



    field = copy.deepcopy(originalfield)

    # schedule field in tournament
    doscheduling(field, nhf)

    # do it until there are no kids left not fully scheduled
    while PlayerClass.failed != 0:
        print "Number of students not able to scheduled: " + str(PlayerClass.failed)
        print "Failed configuration."
        # reset number of kids not fully scheduled
        PlayerClass.failed = 0
        # initialize a new tournament
        del nhf
        nhf = Tournament()
        del field
        field = copy.deepcopy(originalfield)
        # schedule field in tournament
        doscheduling(field, nhf)

    print "Good configuration."


    # generate PDFs
    createpdfs(field)

    # create scoresheets
    createscoresheets(nhf)

    # create master schedules
    createmasters(field, nhf)


if __name__ == "__main__":
    # track time
    t0 = time.time()
    main()
    # calculate length
    t = time.time() - t0
    print "Done."
    print "it took " + str(round(t, 2)) + " to complete."
