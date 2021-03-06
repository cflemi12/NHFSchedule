"""
File for doing work of making the schedule.
@author Chase Fleming
4/23/17

File that includes heavy lifting functions. Does most of the work for scheduling.
"""

from openpyxl import load_workbook, Workbook
from PlayerClass import Player
from FPDFClass import PDF
from shutil import copy2
from operator import itemgetter


def generateplayingfield(info, tournament):
    """
    Generates the playing field given the raw data.
    Gets passed the file name and the tournament.
    """

    # load workbook data
    print "Loading workbook data."
    players = []
    wb = load_workbook(info, read_only=True, data_only=True)
    ws = wb.active

    # construct players from each row value
    for row in ws:
        if str(row[0].value).lower() in ['none']:
            continue
        name = str(row[0].value) + " " + str(row[1].value)
        division = str(row[2].value)
        hometown = str(row[3].value)
        school = str(row[4].value)
        anniversary = str(row[5].value).lower() in ['yes']
        sande = str(row[6].value).lower() in ['yes']
        citizen = str(row[7].value).lower() in ['yes']
        military = str(row[8].value).lower() in ['military']
        geography = str(row[9].value).lower() in ['geography']
        fqn = str(row[10].value).lower in ['yes']
        bowl = str(row[11].value).lower() in ['yes']
        bee = str(row[12].value).lower() in ['yes']
        seed = str(row[13].value).lower()
        newplayer = Player(name, division, hometown, school, bee, bowl, anniversary, sande, citizen,
                           military, geography, fqn, seed, tournament)
        players.append(newplayer)

    # remove header
    players.pop(0)

    # close workbook and return players
    wb.close()

    return players


def createpdfs(players):
    """Creates PDFs."""
    print "Creating PDFs."
    for player in players:
        pdf = PDF()
        pdf.print_schedule(player.name, player.schedule, player.id, player.seed)
        savename = player.name.split(" ")[-1] + ", " + " ".join(player.name.split(" ")[0:-1])
        pdf.output('/Users/chasefleming/PycharmProjects/NHFSchedule/output/Schedules/' + savename + '.pdf', 'F')


def schedulebuz(field, tournament):
    """ Schedules the buzzer portion to the """
    field = list(filter(lambda stu: stu.bee, field))
    seeds = [seed for seed in "abcdefghij"]
    divisions = ['8', '7', "Elementary"]
    friday = tournament.buzzerschedule[0:8]
    saturday = tournament.buzzerschedule[8:]
    eig = [9, 8, 9, 8, 9, 9, 9, 9]
    sev = [8, 8, 8, 8, 8, 8, 8, 8]
    elm = [9, 10, 9, 10, 10, 10, 10, 10]
    divisiontotals = zip(divisions, [eig, sev, elm])

    for div, tots in divisiontotals:
        print "Division: " + div
        for seed in seeds:
            players = list(filter(lambda stu: (stu.division == div) & (stu.seed == seed), field))

            signedup = [0] * 8
            attemptschedule = list(
                map(lambda stu: stu.attemptschedulebuz(signedup, tots, friday), players))
            while not all(ev[0] for ev in attemptschedule):
                signedup = [0] * 8
                attemptschedule = list(
                    map(lambda stu: stu.attemptschedulebuz(signedup, tots, friday), players))
            for attempt, player, schedule in attemptschedule:
                player.updateschedule(schedule)

            attemptschedule = list(
                map(lambda stu: stu.attemptschedulebuz(signedup, tots, saturday), players))
            while not all(ev[0] for ev in attemptschedule):
                signedup = [0] * 8
                attemptschedule = list(
                    map(lambda stu: stu.attemptschedulebuz(signedup, tots, saturday), players))
            for attempt, player, schedule in attemptschedule:
                player.updateschedule(schedule)


def createscoresheets(tournament):
    """ Generates and saves score sheets for events."""
    print "Creating scoresheets."
    seeds = [seed for seed in 'abcdefghij']
    orig = "/Users/chasefleming/PycharmProjects/NHFSchedule/docs/scoresheet.xlsx"

    # create regular buzzer rounds scoresheets
    directory = "/Users/chasefleming/PycharmProjects/NHFSchedule/output/Scoresheets/Buzzer/"
    for i, roundrooms in enumerate(tournament.buzzerrooms):
        for j, room in enumerate(roundrooms):
            if (room.roomnumber + 1) not in tournament.usablerooms:
                continue
            playerspaces = ['B8', 'B9', 'B10', 'B11', 'B12', 'B14', 'B15', 'B16', 'B17', 'B18']
            name = "BuzzerRound" + str(i + 1) + "Room" + str(j + 1) + ".xlsx"
            dst = directory + name
            copy2(orig, dst)
            wb = load_workbook(dst)
            wb.template = True
            ws = wb.active
            ws['M2'] = "NHB Buzzer"
            ws['O2'] = "ACE Room " + str(j + 1)
            ws['O4'] = "Round " + str(i + 1)
            time = room.schedule[room.roundnumber][0][0]
            if time < 200:
                ws['M4'] = "Friday"
            else:
                ws['M4'] = "Saturday"
            if room.roster['a'] is None:
                wb.save(dst)
                wb.close()
            else:
                for space, seed in zip(playerspaces, seeds):
                    stu = room.roster[seed]
                    name = stu.name
                    ws[space] = str(name[0] + "." + name[(name.index(" ")):]) + " " + \
                                str(stu.seed).upper() + "-" + str(stu.id)
                wb.save(dst)
                wb.close()

    # create Citizenship buzzer round scoresheets
    directory = "/Users/chasefleming/PycharmProjects/NHFSchedule/output/Scoresheets/Citizen/"
    for i, roundrooms in enumerate(tournament.citizenrooms):
        for j, room in enumerate(roundrooms):
            if len(room.roster) == 0:
                continue
            playerspaces = ['B8', 'B9', 'B10', 'B11', 'B12', 'B14', 'B15', 'B16', 'B17', 'B18']
            name = "CitizenshipRound" + str(i + 1) + "Room" + str(j + 1) + ".xlsx"
            dst = directory + name
            copy2(orig, dst)
            wb = load_workbook(dst)
            ws = wb.active
            ws['M2'] = "Citzenship"
            ws['O2'] = "Room " + str(j + 1)
            ws['O4'] = "Round " + str(i + 1)
            time = room.schedule[room.roundnumber][0][0]
            if time < 200:
                ws['M4'] = "Friday"
            else:
                ws['M4'] = "Saturday"
            for space in playerspaces:
                ws[space] = ""
            for space, player in zip(playerspaces, room.roster):
                name = player.name
                ws[space] = str(name[0] + "." + name[(name.index(" ")):]) + " " + \
                            str(player.seed).upper() + "-" + str(player.id)
            wb.save(dst)
            wb.close()

    # create Sports and entertainment buzzer round scoresheets
    directory = "/Users/chasefleming/PycharmProjects/NHFSchedule/output/Scoresheets/Sports/"
    for i, roundrooms in enumerate(tournament.sanderooms):
        for j, room in enumerate(roundrooms):
            if len(room.roster) == 0:
                continue
            playerspaces = ['B8', 'B9', 'B10', 'B11', 'B12', 'B14', 'B15', 'B16', 'B17', 'B18']
            name = "S.AndE.Round" + str(i + 1) + "Room" + str(j + 1) + ".xlsx"
            dst = directory + name
            copy2(orig, dst)
            wb = load_workbook(dst)
            ws = wb.active
            ws['M2'] = "S & E"
            ws['O2'] = "ACE Room " + str(j + 1)
            ws['O4'] = "Round " + str(i + 1)
            time = room.schedule[room.roundnumber][0][0]
            if time < 200:
                ws['M4'] = "Friday"
            else:
                ws['M4'] = "Saturday"
            for space in playerspaces:
                ws[space] = ""
            for space, player in zip(playerspaces, room.roster):
                name = player.name
                ws[space] = str(name[0] + "." + name[(name.index(" ")):]) + " " + \
                            str(player.seed).upper() + "-" + str(player.id)
            wb.save(dst)
            wb.close()


def createmasters(field, tournament):
    """ Generates the master schedules. """
    seeds = [seed for seed in 'abcdefghij']
    print "Generating master schedules."
    directory = "/Users/chasefleming/PycharmProjects/NHFSchedule/output/Masters/"

    # do for exams
    sub = "Exams/ExamMaster.xlsx"
    wb = Workbook()
    for rnd in tournament.examrooms:
        ws = wb.create_sheet(title="examround" + str(tournament.examrooms.index(rnd) + 1))
        ws.append(["Exam", "ID", "Name", "Scores"])
        for player in rnd.roster:
            ws.append(["NHB Exam", str(player.seed).upper() + "-" + str(player.id), player.name])
    l = wb.get_sheet_names()[0]
    head = wb.get_sheet_by_name(l)
    wb.remove_sheet(head)
    wb.save(directory + sub)

    # do for military
    sub = "Exams/MilitaryMaster.xlsx"
    wb = Workbook()
    for rnd in tournament.militaryrooms:
        ws = wb.create_sheet(title="militaryexamround" + str(tournament.militaryrooms.index(rnd) + 1))
        ws.append(["Exam", "ID", "Name", "Scores"])
        for player in rnd.roster:
            ws.append(["Military History Subject Exam", str(player.seed).upper() + "-" + str(player.id), player.name])
    l = wb.get_sheet_names()[0]
    head = wb.get_sheet_by_name(l)
    wb.remove_sheet(head)
    wb.save(directory + sub)

    # do for geography
    sub = "Exams/GeographyMaster.xlsx"
    wb = Workbook()
    for rnd in tournament.geographyrooms:
        ws = wb.create_sheet(title="geographysubjectround" + str(tournament.geographyrooms.index(rnd) + 1))
        ws.append(["Exam", "ID", "Name", "Scores"])
        for player in rnd.roster:
            ws.append(["Geogrpahy Subject Exam", str(player.seed).upper() + "-" + str(player.id), player.name])
    l = wb.get_sheet_names()[0]
    head = wb.get_sheet_by_name(l)
    wb.remove_sheet(head)
    wb.save(directory + sub)

    # do buzzer rounds
    sub = "Buzzer/BuzzerMaster.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "ID", "First Round", "Second Round", "Third Round", "Fourth Round", "Exam", "Total"])
    offset = 0
    players = list(filter(lambda stu: stu.bee, field))
    for player in players:
        equations = []
        exam = []
        for event in sorted(player.schedule, key=itemgetter(1)):
            if event[0] == "History Bee Buzzer Round":
                room = str(tournament.buzzerschedule.index(event[1]) + 1)
                eq = "=VLOOKUP(OFFSET(A2," + str(offset) + ",0),$buzzerRound" + room + ".C2:D500, 2, 0)"
                equations.append(eq)
            if event[0] == "History Bee Exam":
                room = str(tournament.examschedule.index(event[1]) + 1)
                eq = "=VLOOKUP(OFFSET(A2," + str(offset) + ",0),$examround" + room + ".C2:D500, 2, 0)"
                exam.append(eq)
        ws.append([player.name, str(player.seed).upper() + "-" + str(player.id), equations[0], equations[1], equations[2], equations[3],
                   exam[0],
                   "=SUM(OFFSET(A2, " + str(offset) + ", 2):OFFSET(A2," + str(offset) + ", 6)) - MIN(OFFSET(A2," +
                   str(offset) + ", 2):OFFSET(A2," + str(offset) + ", 5))"])
        offset += 1
    for rnd in tournament.buzzerrooms:
        ws = wb.create_sheet(title="buzzerround" + str(tournament.buzzerrooms.index(rnd) + 1))
        ws.append(["Room", "ID", "Name", "Score"])
        for room in rnd:
            if room.roster['a'] is None:
                continue
            for s in seeds:
                ros = room.roster
                ws.append([room.roomnumber+1, str(ros[s].seed).upper() + "-" + str(ros[s].id), ros[s].name])
    for rnd in tournament.examrooms:
        ws = wb.create_sheet(title="examround" + str(tournament.examrooms.index(rnd) + 1))
        ws.append(["Exam", "ID", "Name", "Scores"])
        for player in rnd.roster:
            ws.append(["NHB Exam", str(player.seed).upper() + "-" + str(player.id), player.name])
    wb.save(directory + sub)

    # do for citizenship
    sub = "Citizen/CitizenMaster.xlsx"
    wb = Workbook()
    for rnd in tournament.citizenrooms:
        ws = wb.create_sheet(title="CitizenshipBeeRound " + str(tournament.citizenrooms.index(rnd) + 1))
        ws.append(["Room", "ID", "Name", "Score"])
        for room in rnd:
            for player in sorted(room.roster, key=lambda stu: stu.seed):
                ws.append([room.roomnumber+1, str(player.seed).upper() + "-" + str(player.id), player.name])
    l = wb.get_sheet_names()[0]
    head = wb.get_sheet_by_name(l)
    wb.remove_sheet(head)
    wb.save(directory + sub)

    # do for sports and entertainment
    sub = "Sports/SAndEMaster.xlsx"
    wb = Workbook()
    for rnd in tournament.sanderooms:
        ws = wb.create_sheet(title="S.AndeE.BeeRound" + str(tournament.sanderooms.index(rnd) + 1))
        ws.append(["Room", "ID", "Name", "Score"])
        for room in rnd:
            for player in sorted(room.roster, key=lambda stu: stu.seed):
                ws.append([room.roomnumber+1, str(player.seed).upper() + "-" + str(player.id), player.name])
    l = wb.get_sheet_names()[0]
    head = wb.get_sheet_by_name(l)
    wb.remove_sheet(head)
    wb.save(directory + sub)




def doscheduling(field, tournament):
    """Does all the heavy lifting. Makes the schedule for each student."""
    print "Scheduling Military History Subject Exams."
    map(lambda stu: stu.schedulemil(tournament), field)
    print "Scheduling Geography Subject Exams."
    map(lambda stu: stu.schedulegeo(tournament), field)
    print "Scheduling Side Events."
    map(lambda stu: stu.schedulecit(tournament), field)
    map(lambda stu: stu.schedulesae(tournament), field)
    print "Scheduling Buzzer rounds for..."
    schedulebuz(field, tournament)
    print "Scheduling Exams."
    map(lambda stu: stu.scheduleexm(tournament), field)

    print "Setting Exam Rooms."
    tournament.scheduleexamrooms(field)
    print "Setting Side Event Rooms."
    tournament.schedulesiderooms(field)
    print "Setting Buzzer Rooms."
    tournament.schedulebuzzerrooms(field)
